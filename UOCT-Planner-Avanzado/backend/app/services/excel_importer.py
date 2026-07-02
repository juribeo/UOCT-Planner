from openpyxl import load_workbook
from app.models.entities import Solicitud, Historial
from app.services.helpers import pick

def import_excel(path, db):
    wb=load_workbook(path, data_only=True); ws=wb.active; rows=list(ws.iter_rows(values_only=True))
    if not rows: return 0
    headers=[str(h).strip() if h is not None else '' for h in rows[0]]; count=0
    for values in rows[1:]:
        data={}
        for i,h in enumerate(headers):
            if h: data[h]='' if i>=len(values) or values[i] is None else str(values[i]).strip()
        empresa=pick(data,['Empresa','Contratista','Solicitante'])
        proyecto=pick(data,['Proyecto','Nombre Proyecto','Cruce','Intersección','Interseccion'])
        municipio=pick(data,['Municipio','Municipalidad'])
        comuna=pick(data,['Comuna'])
        direccion=pick(data,['Dirección','Direccion','Ubicación','Ubicacion','Cruce','Intersección','Interseccion'])
        tipo=pick(data,['Tipo Actividad','Actividad','Tipo de actividad','Solicitud','Tipo'])
        fecha=pick(data,['Fecha','Fecha solicitada','Fecha Actividad','Programación','Programacion'])
        obs=pick(data,['Observaciones','Observación','Comentario','Comentarios'])
        contacto=pick(data,['Contacto','Nombre contacto']); correo=pick(data,['Correo','Email','Correo contacto'])
        if not any([empresa,proyecto,comuna,direccion,tipo,obs]): continue
        sol=Solicitud(empresa=empresa,proyecto=proyecto,municipio=municipio,comuna=comuna,direccion=direccion,tipo_actividad=tipo or 'Solicitud',fecha_solicitada=fecha,observaciones=obs,contacto=contacto,correo_contacto=correo,origen='Excel')
        db.add(sol); db.flush(); db.add(Historial(solicitud_id=sol.id,accion='Importación Excel',detalle='Solicitud creada desde Excel',usuario='Sistema')); count+=1
    db.commit(); return count
